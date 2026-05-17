"""
Excel (xlsx) generators for the three admin reports.

Presentation overhaul: merged titles with clean hierarchy, auto-sized
columns, navy header band, zebra body, distinct total row, and ₹
currency formatting so Excel renders amounts like ₹1,92,749.00.

Data contracts and calculations are unchanged from the reporting
modules. This module only controls visual presentation.
"""
from __future__ import annotations

import datetime
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.services.pdf_service import COMPANY_NAME

# ── Palette (matches PDF navy/grey scheme) ───────────────────────────────────

_NAVY = "1F3A5F"
_HEADER_FG = "FFFFFF"
_STRIPE = "F7F9FC"
_TOTAL_BG = "E8ECF3"
_RULE = "C8CED9"
_RULE_DARK = "3A4A63"
_TEXT = "1F2A3A"
_MUTED = "6B7280"


def _fonts():
    return {
        "company": Font(name="Calibri", size=16, bold=True, color=_TEXT),
        "route": Font(name="Calibri", size=11, color=_MUTED),
        "title": Font(name="Calibri", size=13, bold=True, color=_TEXT),
        "subtitle": Font(name="Calibri", size=10, italic=True, color=_MUTED),
        "header": Font(name="Calibri", size=11, bold=True, color=_HEADER_FG),
        "body": Font(name="Calibri", size=11, color=_TEXT),
        "total": Font(name="Calibri", size=11, bold=True, color=_TEXT),
        "grand": Font(name="Calibri", size=12, bold=True, color=_TEXT),
        "summary_label": Font(name="Calibri", size=11, bold=True, color=_TEXT),
    }


_CENTER = Alignment(horizontal="center", vertical="center")
_LEFT = Alignment(horizontal="left", vertical="center", indent=1)
_RIGHT = Alignment(horizontal="right", vertical="center", indent=1)

_HEADER_FILL = PatternFill("solid", fgColor=_NAVY)
_TOTAL_FILL = PatternFill("solid", fgColor=_TOTAL_BG)
_STRIPE_FILL = PatternFill("solid", fgColor=_STRIPE)

_BORDER_TOP_DARK = Border(top=Side(border_style="medium", color=_RULE_DARK))
_BORDER_BOTTOM_LIGHT = Border(bottom=Side(border_style="thin", color=_RULE))
_BORDER_HEADER_BOTTOM = Border(bottom=Side(border_style="medium", color=_RULE_DARK))

# Indian locale (0x4009 = en-IN) + Indian digit grouping "#,##,##0".
# Renders 192749 → ₹1,92,749.00, not the Western ₹192,749.00.
_CUR_FMT = '[$₹-4009] #,##,##0.00;[Red]-[$₹-4009] #,##,##0.00'
_RATE_FMT = '[$₹-4009] #,##0.00'   # small rates never need lakh grouping
_INT_FMT = '#,##,##0'
_DATE_FMT = "dd-mmm-yyyy"


# ── Helpers ──────────────────────────────────────────────────────────────────


def _to_number(val) -> float:
    """Convert a decimal/string amount to float for Excel numeric cells."""
    if val is None or val == "":
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, Decimal):
        return float(val)
    try:
        return float(str(val).replace(",", "").replace("₹", "").strip())
    except (ValueError, TypeError):
        return 0.0


def _autosize(ws: Worksheet, hints: list[int], data_rows: list[list] | None = None) -> None:
    """Set column widths. `hints` is the minimum width per column; if data_rows
    is given, widen up to the longest value."""
    widths = list(hints)
    if data_rows:
        for row in data_rows:
            for i, cell in enumerate(row):
                if i >= len(widths):
                    break
                length = len(str(cell)) if cell is not None else 0
                widths[i] = max(widths[i], min(length + 2, 50))
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_title_block(
    ws: Worksheet, data: dict, title: str, col_span: int
) -> int:
    """Write company → route → title → subtitle rows (all merged). Returns
    the next row number to use for the table header."""
    f = _fonts()

    ws.row_dimensions[1].height = 22
    c = ws.cell(row=1, column=1, value=COMPANY_NAME)
    c.font = f["company"]
    c.alignment = _CENTER
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=col_span)

    c = ws.cell(row=2, column=1, value=data["route_label"])
    c.font = f["route"]
    c.alignment = _CENTER
    ws.merge_cells(start_row=2, end_row=2, start_column=1, end_column=col_span)

    ws.row_dimensions[3].height = 18
    c = ws.cell(row=3, column=1, value=title)
    c.font = f["title"]
    c.alignment = _CENTER
    ws.merge_cells(start_row=3, end_row=3, start_column=1, end_column=col_span)

    dfrom = data["date_from"]
    dto = data["date_to"]
    if isinstance(dfrom, datetime.date):
        subtitle = f"Period: {dfrom:%d %b %Y}  to  {dto:%d %b %Y}"
    else:
        subtitle = f"Period: {dfrom} to {dto}"
    c = ws.cell(row=4, column=1, value=subtitle)
    c.font = f["subtitle"]
    c.alignment = _CENTER
    ws.merge_cells(start_row=4, end_row=4, start_column=1, end_column=col_span)

    return 6  # leave row 5 empty as a spacer


def _style_header_row(ws: Worksheet, row: int, cols: int) -> None:
    ws.row_dimensions[row].height = 22
    f = _fonts()
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = _HEADER_FILL
        cell.font = f["header"]
        cell.alignment = _CENTER
        cell.border = _BORDER_HEADER_BOTTOM


def _apply_zebra(ws: Worksheet, first_row: int, last_row: int, cols: int) -> None:
    for r in range(first_row, last_row + 1):
        if (r - first_row) % 2 == 1:  # second, fourth, … row
            for c in range(1, cols + 1):
                ws.cell(row=r, column=c).fill = _STRIPE_FILL


def _style_total_row(ws: Worksheet, row: int, cols: int) -> None:
    ws.row_dimensions[row].height = 22
    f = _fonts()
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = _TOTAL_FILL
        cell.font = f["total"]
        cell.border = _BORDER_TOP_DARK


# ── Report A: Itemwise Levy Summary ──────────────────────────────────────────


def generate_itemwise_levy_xlsx(data: dict) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Itemwise Levy"
    ws.sheet_view.showGridLines = False
    f = _fonts()

    branches = data["branches"]
    col_span = 4 + len(branches)  # Item, Levy, <branches>, Quantity, Amount
    next_row = _write_title_block(ws, data, "Itemwise Levy Summary", col_span)

    # Header
    header = ["Item", "Levy"] + [b["name"] for b in branches] + ["Quantity", "Amount"]
    for i, h in enumerate(header, start=1):
        ws.cell(row=next_row, column=i, value=h)
    _style_header_row(ws, next_row, col_span)
    header_row = next_row
    first_body = next_row + 1
    next_row = first_body

    # Body
    for row in data["rows"]:
        ws.cell(row=next_row, column=1, value=row["item_name"]).alignment = _LEFT
        c = ws.cell(row=next_row, column=2, value=_to_number(row["levy"]))
        c.number_format = _RATE_FMT
        c.alignment = _RIGHT
        for j, b in enumerate(branches, start=3):
            c = ws.cell(
                row=next_row,
                column=j,
                value=int(row["branch_quantities"].get(str(b["id"]), 0)),
            )
            c.number_format = _INT_FMT
            c.alignment = _RIGHT
        c = ws.cell(row=next_row, column=3 + len(branches), value=int(row["total_quantity"]))
        c.number_format = _INT_FMT
        c.alignment = _RIGHT
        c = ws.cell(row=next_row, column=4 + len(branches), value=_to_number(row["amount"]))
        c.number_format = _CUR_FMT
        c.alignment = _RIGHT
        ws.cell(row=next_row, column=1).font = f["body"]
        next_row += 1

    last_body = next_row - 1
    if last_body >= first_body:
        _apply_zebra(ws, first_body, last_body, col_span)

    # Total row
    ws.cell(row=next_row, column=1, value="Total").alignment = _LEFT
    c = ws.cell(row=next_row, column=col_span, value=_to_number(data["grand_total"]))
    c.number_format = _CUR_FMT
    c.alignment = _RIGHT
    _style_total_row(ws, next_row, col_span)
    next_row += 2

    # Summary block
    c = ws.cell(row=next_row, column=1, value="Summary")
    c.font = f["summary_label"]
    next_row += 1
    summary_start = next_row
    for b in branches:
        ws.cell(row=next_row, column=1, value=b["name"]).font = f["body"]
        ws.cell(row=next_row, column=1).alignment = _LEFT
        amt = ws.cell(
            row=next_row, column=2,
            value=_to_number(data["branch_totals"].get(str(b["id"]), "0")),
        )
        amt.number_format = _CUR_FMT
        amt.alignment = _RIGHT
        amt.font = f["body"]
        next_row += 1
    # Grand total line
    ws.cell(row=next_row, column=1, value="Grand Total").font = f["grand"]
    ws.cell(row=next_row, column=1).alignment = _LEFT
    gt = ws.cell(row=next_row, column=2, value=_to_number(data["grand_total"]))
    gt.number_format = _CUR_FMT
    gt.alignment = _RIGHT
    gt.font = f["grand"]
    ws.cell(row=next_row, column=1).border = _BORDER_TOP_DARK
    ws.cell(row=next_row, column=2).border = _BORDER_TOP_DARK

    # Freeze header
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate

    # Sizes
    data_preview = [header] + [
        [r["item_name"]] + [""] * (col_span - 1) for r in data["rows"]
    ]
    _autosize(ws, [34, 12] + [15] * len(branches) + [14, 18], data_rows=data_preview)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Report B: Date-Wise Branch Summary ───────────────────────────────────────


def generate_date_branch_summary_xlsx(data: dict) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Date-Branch Summary"
    ws.sheet_view.showGridLines = False
    f = _fonts()

    cols = data["columns"]
    col_span = 2 + len(cols)  # Date + mode cols + Total
    next_row = _write_title_block(
        ws, data, "Date-Wise Branch Summary  —  Cash & UPI", col_span
    )

    header = ["Date"] + [c["label"] for c in cols] + ["Total"]
    for i, h in enumerate(header, start=1):
        ws.cell(row=next_row, column=i, value=h)
    _style_header_row(ws, next_row, col_span)
    header_row = next_row
    first_body = next_row + 1
    next_row = first_body

    for row in data["rows"]:
        d = row["date"]
        dc = ws.cell(
            row=next_row, column=1,
            value=d if isinstance(d, datetime.date) else str(d),
        )
        if isinstance(d, datetime.date):
            dc.number_format = _DATE_FMT
        dc.alignment = _LEFT
        dc.font = f["body"]

        for j, c in enumerate(cols, start=2):
            val = _to_number(row["cells"].get(c["key"], "0"))
            cell = ws.cell(row=next_row, column=j, value=val if val else None)
            cell.number_format = _CUR_FMT
            cell.alignment = _RIGHT
            cell.font = f["body"]

        tot = ws.cell(row=next_row, column=col_span, value=_to_number(row["total"]))
        tot.number_format = _CUR_FMT
        tot.alignment = _RIGHT
        tot.font = f["body"]
        next_row += 1

    last_body = next_row - 1
    if last_body >= first_body:
        _apply_zebra(ws, first_body, last_body, col_span)

    # Total row
    ws.cell(row=next_row, column=1, value="Total").alignment = _LEFT
    for j, c in enumerate(cols, start=2):
        cell = ws.cell(
            row=next_row, column=j,
            value=_to_number(data["column_totals"].get(c["key"], "0")),
        )
        cell.number_format = _CUR_FMT
        cell.alignment = _RIGHT
    gt = ws.cell(row=next_row, column=col_span, value=_to_number(data["grand_total"]))
    gt.number_format = _CUR_FMT
    gt.alignment = _RIGHT
    _style_total_row(ws, next_row, col_span)

    ws.freeze_panes = ws.cell(row=header_row + 1, column=2).coordinate
    _autosize(ws, [14] + [18] * len(cols) + [18])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Report C: Itemwise Daily Collection Charges ──────────────────────────────


def generate_itemwise_daily_charges_xlsx(data: dict) -> BytesIO:
    """Flat sheet with Date / Branch grouping + a Summary sheet."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Charges"
    ws.sheet_view.showGridLines = False
    f = _fonts()

    col_span = 7  # Date | Branch | Item | Charges | Quantity | Levy | Amount
    next_row = _write_title_block(
        ws, data, "Itemwise Daily Collection Charges Summary", col_span
    )

    header = ["Date", "Branch", "Item", "Charges", "Quantity", "Levy", "Amount"]
    for i, h in enumerate(header, start=1):
        ws.cell(row=next_row, column=i, value=h)
    _style_header_row(ws, next_row, col_span)
    header_row = next_row
    first_body = next_row + 1
    next_row = first_body

    for ds in data["dates"]:
        d = ds["date"]
        section_start = next_row
        for bs in ds["branches"]:
            for r in bs["rows"]:
                dc = ws.cell(
                    row=next_row, column=1,
                    value=d if isinstance(d, datetime.date) else str(d),
                )
                if isinstance(d, datetime.date):
                    dc.number_format = _DATE_FMT
                dc.alignment = _LEFT
                dc.font = f["body"]

                ws.cell(row=next_row, column=2, value=bs["branch_name"]).alignment = _LEFT
                ws.cell(row=next_row, column=2).font = f["body"]

                ws.cell(row=next_row, column=3, value=r["item_name"]).alignment = _LEFT
                ws.cell(row=next_row, column=3).font = f["body"]

                ch = ws.cell(row=next_row, column=4, value=_to_number(r["charges"]))
                ch.number_format = _RATE_FMT
                ch.alignment = _RIGHT

                q = ws.cell(row=next_row, column=5, value=int(r["quantity"]))
                q.number_format = _INT_FMT
                q.alignment = _RIGHT

                lv = ws.cell(row=next_row, column=6, value=_to_number(r["levy"]))
                lv.number_format = _CUR_FMT
                lv.alignment = _RIGHT

                am = ws.cell(row=next_row, column=7, value=_to_number(r["amount"]))
                am.number_format = _CUR_FMT
                am.alignment = _RIGHT
                next_row += 1

            # Branch subtotal line — subtotal = sum(Ticket.net_amount), shown
            # in the Amount column; item Amount cells may sum higher if a
            # ticket carried a discount.
            ws.cell(
                row=next_row, column=3, value=f"{bs['branch_name']} subtotal"
            ).font = f["total"]
            sub = ws.cell(row=next_row, column=7, value=_to_number(bs["subtotal"]))
            sub.number_format = _CUR_FMT
            sub.alignment = _RIGHT
            sub.font = f["total"]
            ws.cell(row=next_row, column=7).border = _BORDER_TOP_DARK
            next_row += 1

        # Day total strip (merged Date+Branch+Item+Charges+Quantity+Levy columns)
        label = ws.cell(
            row=next_row, column=1,
            value=f"Total for {d:%d %b %Y}" if isinstance(d, datetime.date) else f"Total for {d}",
        )
        label.font = f["total"]
        label.alignment = _LEFT
        ws.merge_cells(start_row=next_row, end_row=next_row, start_column=1, end_column=6)
        day_total = ws.cell(row=next_row, column=7, value=_to_number(ds["day_total"]))
        day_total.number_format = _CUR_FMT
        day_total.alignment = _RIGHT
        _style_total_row(ws, next_row, col_span)
        next_row += 2  # blank line after each day

    # Grand total
    label = ws.cell(row=next_row, column=1, value="GRAND TOTAL")
    label.font = f["grand"]
    label.alignment = _LEFT
    ws.merge_cells(start_row=next_row, end_row=next_row, start_column=1, end_column=6)
    gt = ws.cell(row=next_row, column=7, value=_to_number(data["grand_total"]))
    gt.number_format = _CUR_FMT
    gt.alignment = _RIGHT
    gt.font = f["grand"]
    ws.row_dimensions[next_row].height = 26
    for c in range(1, col_span + 1):
        cell = ws.cell(row=next_row, column=c)
        cell.fill = _TOTAL_FILL
        cell.border = Border(
            top=Side(border_style="medium", color=_RULE_DARK),
            bottom=Side(border_style="medium", color=_RULE_DARK),
        )

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate
    _autosize(ws, [14, 20, 36, 14, 14, 14, 18])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Report D: Month-Wise Branch Summary ──────────────────────────────────────


def generate_month_branch_summary_xlsx(data: dict) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Month-Branch Summary"
    ws.sheet_view.showGridLines = False
    f = _fonts()

    cols = data["columns"]
    col_span = 2 + len(cols)  # Month + mode cols + Total
    next_row = _write_title_block(
        ws, data, "Month-Wise Branch Summary  —  Cash & UPI", col_span
    )

    header = ["Month"] + [c["label"] for c in cols] + ["Total"]
    for i, h in enumerate(header, start=1):
        ws.cell(row=next_row, column=i, value=h)
    _style_header_row(ws, next_row, col_span)
    header_row = next_row
    first_body = next_row + 1
    next_row = first_body

    for row in data["rows"]:
        # Month label is "MM-YYYY" — keep as plain text, not a date object,
        # because Excel's date format would mis-render it as a day.
        mc = ws.cell(row=next_row, column=1, value=row["month_label"])
        mc.alignment = _LEFT
        mc.font = f["body"]

        for j, c in enumerate(cols, start=2):
            val = _to_number(row["cells"].get(c["key"], "0"))
            cell = ws.cell(row=next_row, column=j, value=val if val else None)
            cell.number_format = _CUR_FMT
            cell.alignment = _RIGHT
            cell.font = f["body"]

        tot = ws.cell(row=next_row, column=col_span, value=_to_number(row["total"]))
        tot.number_format = _CUR_FMT
        tot.alignment = _RIGHT
        tot.font = f["body"]
        next_row += 1

    last_body = next_row - 1
    if last_body >= first_body:
        _apply_zebra(ws, first_body, last_body, col_span)

    # Total row
    ws.cell(row=next_row, column=1, value="Total").alignment = _LEFT
    for j, c in enumerate(cols, start=2):
        cell = ws.cell(
            row=next_row, column=j,
            value=_to_number(data["column_totals"].get(c["key"], "0")),
        )
        cell.number_format = _CUR_FMT
        cell.alignment = _RIGHT
    gt = ws.cell(row=next_row, column=col_span, value=_to_number(data["grand_total"]))
    gt.number_format = _CUR_FMT
    gt.alignment = _RIGHT
    _style_total_row(ws, next_row, col_span)

    ws.freeze_panes = ws.cell(row=header_row + 1, column=2).coordinate
    _autosize(ws, [12] + [16] * len(cols) + [18])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
