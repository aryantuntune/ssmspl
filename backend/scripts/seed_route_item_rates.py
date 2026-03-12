#!/usr/bin/env python3
"""
Seed Route Item Rates from Excel Dataset
=========================================

Reads item rates per route from the Excel file and upserts into the
item_rates table.  Idempotent — safe to run multiple times.

Usage:
    python scripts/seed_route_item_rates.py                # execute changes
    python scripts/seed_route_item_rates.py --dry-run       # preview only
    python scripts/seed_route_item_rates.py --env .env.production  # custom env

Excel path: data/item_rates/item_rates_list_final_all_routes.xlsx
"""

from __future__ import annotations

import argparse
import asyncio
import io
import os
import sys
from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path

# Force UTF-8 stdout on Windows (Devanagari text in item names)
if sys.stdout.encoding != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

import openpyxl

# ---------------------------------------------------------------------------
# Resolve project paths
# ---------------------------------------------------------------------------
SCRIPT_DIR = Path(__file__).resolve().parent          # backend/scripts/
BACKEND_DIR = SCRIPT_DIR.parent                       # backend/
PROJECT_DIR = BACKEND_DIR.parent                      # project root

DEFAULT_EXCEL = PROJECT_DIR / "data" / "item_rates" / "item_rates_list_final_all_routes.xlsx"
DEFAULT_ENV   = BACKEND_DIR / ".env.development"

# ---------------------------------------------------------------------------
# Excel sheet name  →  DB route_id
# ---------------------------------------------------------------------------
ROUTE_SHEET_MAP: dict[str, int] = {
    "DABHOL DHOPAVE":    1,   # DABHOL (101) <-> DHOPAVE (102)
    "VESAVI BAGMANDALE": 2,   # VESHVI (103) <-> BAGMANDALE (104)
    "JAIGAD TAVSAL":     3,   # JAIGAD (105) <-> TAVSAL (106)
    "DIGHI AGARDANDA":   4,   # AGARDANDA (107) <-> DIGHI (108)
    "VASAI BHAYANDER":   5,   # BHAYANDER (110) <-> VASAI (109)
    "VIRAR SAPHALE":     7,   # VIRAR (113) <-> SAFALE (114)
}

# Route 6 (AMBET <-> MHAPRAL) is NOT in the Excel — left untouched.

# ---------------------------------------------------------------------------
# Master item number  →  DB item_id
#
# The "master item number" is the canonical serial from Sheet1 in the Excel.
# ---------------------------------------------------------------------------
MASTER_ITEM_MAP: dict[int, int] = {
    1:  11,   # Passenger Adult Above 12 Yr  (प्रवासी प्रौढ)
    2:  12,   # Passenger Child 3-12 Yr      (प्रवासी लहान)
    3:   2,   # Motorcycle With Driver        (मोटारसायकल)
    4:   7,   # Empty Car 5-Str Hatchback     (रिकामी कार hatch back)
    5:   8,   # Empty Luxury Car 5-Str Sedan  (रिकामी लक्झरी कार sedan)
    6:   4,   # Empty 3-Wheeler Rickshaw      (तीन चाकी रिक्षा)
    7:  15,   # 407 Tempo                     (टाटा ४०७)
    8:  18,   # Med Goods 6-Wheeler 709       (टाटा ७०९ / आयशर १०९५)
    9:  23,   # Goods Per Half Ton            (माल प्रति अर्धा टन)
    10: 21,   # Passenger Bus / Truck / Tanker (पॅसेंजर बस / ट्रक / ट्रॅक्टर)
    11: 32,   # 10-Wheeler Truck / JCB        (दहाचाकी मालवाहू ट्रक / जे.सी.बी.)
    12: 33,   # Tractor With Trolley          (ट्रॅक्टर ट्रॉली)
    13:  1,   # Cycle                         (सायकल)
    14: 31,   # Fish / Chicken / Birds / Fruits (मासे / पक्षी / फळे)
    15: 26,   # Cow / Buffalo                 (गाय / बैल / म्हैस)
    16: 27,   # Student Month Pass Up To 7th  (विद्यार्थी मासिक पास पर्यंत)
    17: 28,   # Student Month Pass Above 7th  (विद्यार्थी मासिक पास नंतर)
    18: 30,   # Monthly Pass Passenger        (प्रवासी मासिक पास)
}

# English labels for display
ENGLISH_NAMES: dict[int, str] = {
    1:  "Passenger Adult (प्रवासी प्रौढ)",
    2:  "Passenger Child (प्रवासी लहान)",
    3:  "Motorcycle (मोटारसायकल)",
    4:  "Car Hatchback (रिकामी कार)",
    5:  "Luxury Car Sedan (रिकामी लक्झरी कार)",
    6:  "3-Wheeler Rickshaw (तीन चाकी रिक्षा)",
    7:  "407 Tempo (टाटा ४०७)",
    8:  "6-Wheeler 709 (टाटा ७०९)",
    9:  "Goods Per Half Ton (माल प्रति अर्धा टन)",
    10: "Bus / Truck / Tanker (पॅसेंजर बस)",
    11: "10-Wheeler Truck / JCB (दहाचाकी)",
    12: "Tractor With Trolley (ट्रॅक्टर ट्रॉली)",
    13: "Cycle (सायकल)",
    14: "Fish / Birds / Fruits (मासे / पक्षी / फळे)",
    15: "Cow / Buffalo (गाय / बैल / म्हैस)",
    16: "Student Pass Up To 7th Std",
    17: "Student Pass Above 7th Std",
    18: "Monthly Pass Passenger (प्रवासी मासिक पास)",
}


# ---------------------------------------------------------------------------
# Marathi name → master item number identification
# ---------------------------------------------------------------------------
def identify_master_item(marathi_name: str) -> list[int]:
    """Return master item number(s) detected from the Marathi column B text.

    For the VASAI combined car+luxury row, returns [4, 5].
    """
    name = marathi_name.strip()

    # Combined hatchback + luxury (VASAI special case)
    if "hatch" in name.lower() and "लक्झरी" in name:
        return [4, 5]

    # Monthly passes — check BEFORE generic "प्रवासी" match
    if "प्रवासी मासिक" in name:
        return [18]
    if "विद्यार्थी" in name and "पर्यंत" in name:
        return [16]
    if "विद्यार्थी" in name and "नंतर" in name:
        return [17]

    # Passengers
    if "प्रौढ" in name:
        return [1]
    if "लहान" in name:
        return [2]

    # Vehicles
    if "मोटारसायकल" in name:
        return [3]
    if "hatch" in name.lower():
        return [4]
    if "लक्झरी" in name or "sedan" in name.lower():
        return [5]
    if "तीन चाकी" in name or "रिक्षा" in name:
        return [6]
    if "४०७" in name:
        return [7]
    if "७०९" in name:
        return [8]

    # Goods
    if "माल" in name and "अर्धा" in name:
        return [9]

    # Heavy vehicles
    if "पॅसेंजर बस" in name:
        return [10]
    if "दहाचाकी" in name:
        return [11]
    if "ट्रॅkली" in name or "ट्रॉली" in name:
        return [12]

    # Other
    if "सायकल" in name:
        return [13]
    if "मासे" in name or "पक्षी" in name:
        return [14]
    if "गाय" in name:
        return [15]

    return []


# ---------------------------------------------------------------------------
# Data structures
# ---------------------------------------------------------------------------
@dataclass
class RateEntry:
    """One item-rate row to upsert."""
    route_id: int
    item_id: int
    rate: Decimal
    levy: Decimal
    master_num: int      # for display
    english_name: str    # for display


@dataclass
class RouteResult:
    """Summary for one route."""
    route_name: str
    added: int = 0
    updated: int = 0
    skipped: int = 0
    errors: list[str] | None = None

    def __post_init__(self) -> None:
        if self.errors is None:
            self.errors = []


# ---------------------------------------------------------------------------
# Excel reader
# ---------------------------------------------------------------------------
def read_excel(excel_path: Path) -> list[RateEntry]:
    """Parse the Excel workbook and return a flat list of RateEntry items."""
    wb = openpyxl.load_workbook(str(excel_path), data_only=True)
    entries: list[RateEntry] = []

    for sheet_name, route_id in ROUTE_SHEET_MAP.items():
        if sheet_name not in wb.sheetnames:
            print(f"  WARNING: Sheet '{sheet_name}' not found in workbook — skipped")
            continue

        ws = wb[sheet_name]
        # Data rows start at row 6 (rows 1-5 are headers)
        for row in ws.iter_rows(min_row=6, max_row=ws.max_row, max_col=5, values_only=True):
            serial, marathi_name, rate_val, levy_val, _total = (
                row[0], row[1], row[2], row[3], row[4] if len(row) > 4 else None
            )

            # Skip empty / footer rows
            if serial is None or marathi_name is None or rate_val is None:
                continue

            master_nums = identify_master_item(str(marathi_name))
            if not master_nums:
                print(f"  WARNING: Could not identify item in '{sheet_name}' "
                      f"row serial={serial}: {marathi_name!r}")
                continue

            rate = Decimal(str(rate_val))
            levy = Decimal(str(levy_val)) if levy_val is not None else Decimal("0")

            for mnum in master_nums:
                item_id = MASTER_ITEM_MAP.get(mnum)
                if item_id is None:
                    print(f"  WARNING: Master item #{mnum} has no DB item_id mapping")
                    continue
                entries.append(RateEntry(
                    route_id=route_id,
                    item_id=item_id,
                    rate=rate,
                    levy=levy,
                    master_num=mnum,
                    english_name=ENGLISH_NAMES.get(mnum, f"Item #{mnum}"),
                ))

    wb.close()
    return entries


# ---------------------------------------------------------------------------
# Database operations (asyncpg)
# ---------------------------------------------------------------------------
def load_database_url(env_path: Path) -> str:
    """Read DATABASE_URL from the env file, convert to asyncpg-compatible URL."""
    if not env_path.exists():
        sys.exit(f"ERROR: Env file not found: {env_path}")

    with open(env_path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line.startswith("#") or "=" not in line:
                continue
            key, _, val = line.partition("=")
            if key.strip() == "DATABASE_URL":
                url = val.strip().strip("'\"")
                # Convert SQLAlchemy asyncpg URL to plain PostgreSQL URL
                url = url.replace("postgresql+asyncpg://", "postgresql://")
                return url

    sys.exit(f"ERROR: DATABASE_URL not found in {env_path}")


async def upsert_rates(
    entries: list[RateEntry],
    db_url: str,
    dry_run: bool,
) -> list[RouteResult]:
    """Connect to database and upsert item rates. Returns per-route results."""
    import asyncpg  # import here so --help works without asyncpg

    conn = await asyncpg.connect(db_url)
    try:
        # Group entries by route_id for per-route reporting
        routes_seen: dict[int, RouteResult] = {}
        route_name_map = {v: k for k, v in ROUTE_SHEET_MAP.items()}

        for entry in entries:
            route_id = entry.route_id
            if route_id not in routes_seen:
                routes_seen[route_id] = RouteResult(
                    route_name=route_name_map.get(route_id, f"Route {route_id}")
                )
            result = routes_seen[route_id]

            # Check if item_rate already exists for this (item_id, route_id)
            existing = await conn.fetchrow(
                "SELECT id, rate, levy, is_active FROM item_rates "
                "WHERE item_id = $1 AND route_id = $2",
                entry.item_id, entry.route_id,
            )

            if existing is None:
                # INSERT new item_rate
                if not dry_run:
                    await conn.execute(
                        "INSERT INTO item_rates (levy, rate, item_id, route_id, is_active, created_at) "
                        "VALUES ($1, $2, $3, $4, TRUE, NOW())",
                        float(entry.levy), float(entry.rate),
                        entry.item_id, entry.route_id,
                    )
                result.added += 1
                print(f"  [ADD]  {result.route_name} | {entry.english_name} | "
                      f"rate={entry.rate} levy={entry.levy}"
                      f"{' (DRY RUN)' if dry_run else ''}")

            else:
                existing_rate = Decimal(str(existing["rate"])) if existing["rate"] is not None else None
                existing_levy = Decimal(str(existing["levy"])) if existing["levy"] is not None else None
                needs_activate = not existing["is_active"]

                rate_changed = existing_rate != entry.rate
                levy_changed = existing_levy != entry.levy

                if rate_changed or levy_changed or needs_activate:
                    if not dry_run:
                        await conn.execute(
                            "UPDATE item_rates SET rate = $1, levy = $2, is_active = TRUE, "
                            "updated_at = NOW() WHERE id = $3",
                            float(entry.rate), float(entry.levy), existing["id"],
                        )
                    result.updated += 1
                    changes = []
                    if rate_changed:
                        changes.append(f"rate: {existing_rate}→{entry.rate}")
                    if levy_changed:
                        changes.append(f"levy: {existing_levy}→{entry.levy}")
                    if needs_activate:
                        changes.append("reactivated")
                    print(f"  [UPD]  {result.route_name} | {entry.english_name} | "
                          f"{', '.join(changes)}"
                          f"{' (DRY RUN)' if dry_run else ''}")
                else:
                    result.skipped += 1

        return list(routes_seen.values())
    finally:
        await conn.close()


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
async def async_main(args: argparse.Namespace) -> None:
    excel_path = Path(args.excel)
    env_path = Path(args.env)

    if not excel_path.exists():
        sys.exit(f"ERROR: Excel file not found: {excel_path}")

    mode = "DRY RUN" if args.dry_run else "LIVE"
    print(f"\n{'='*60}")
    print(f"  Seed Route Item Rates  [{mode}]")
    print(f"{'='*60}")
    print(f"  Excel : {excel_path}")
    print(f"  Env   : {env_path}")
    print()

    # Step 1: Read Excel
    print("Reading Excel file...")
    entries = read_excel(excel_path)
    print(f"  Parsed {len(entries)} item-rate entries from {len(ROUTE_SHEET_MAP)} route sheets\n")

    if not entries:
        print("No entries to process. Exiting.")
        return

    # Step 2: Validate — check no duplicate (item_id, route_id) pairs
    seen_pairs: set[tuple[int, int]] = set()
    for e in entries:
        pair = (e.item_id, e.route_id)
        if pair in seen_pairs:
            print(f"  WARNING: Duplicate entry for item_id={e.item_id} route_id={e.route_id}")
        seen_pairs.add(pair)

    # Step 3: Database upsert
    db_url = load_database_url(env_path)
    print("Connecting to database...")
    results = await upsert_rates(entries, db_url, args.dry_run)

    # Step 4: Per-route summary
    print(f"\n{'='*60}")
    print("  Per-Route Summary")
    print(f"{'='*60}")
    total_added = total_updated = total_skipped = 0
    for r in results:
        print(f"  Route: {r.route_name}")
        print(f"    Added:   {r.added}")
        print(f"    Updated: {r.updated}")
        print(f"    Skipped: {r.skipped}")
        total_added += r.added
        total_updated += r.updated
        total_skipped += r.skipped
        if r.errors:
            for err in r.errors:
                print(f"    ERROR: {err}")

    # Step 5: Grand summary
    print(f"\n{'='*60}")
    print("  Grand Summary")
    print(f"{'='*60}")
    print(f"  Routes processed : {len(results)}")
    print(f"  Total items added: {total_added}")
    print(f"  Total updated    : {total_updated}")
    print(f"  Total skipped    : {total_skipped}")
    if args.dry_run:
        print(f"\n  ** DRY RUN — no changes were written to the database **")
    print()


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Seed route-wise item rates from Excel dataset into the database."
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Preview changes without writing to the database.",
    )
    parser.add_argument(
        "--excel",
        default=str(DEFAULT_EXCEL),
        help=f"Path to the Excel file (default: {DEFAULT_EXCEL})",
    )
    parser.add_argument(
        "--env",
        default=str(DEFAULT_ENV),
        help=f"Path to the .env file with DATABASE_URL (default: {DEFAULT_ENV})",
    )
    args = parser.parse_args()
    asyncio.run(async_main(args))


if __name__ == "__main__":
    main()
